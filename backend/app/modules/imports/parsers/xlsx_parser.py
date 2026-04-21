"""XLSX parser for Clear/XP corretora position snapshots and XP account extracts.

Supports three formats (auto-detected):
  1. Clear PosicaoDetalhada.xlsx  — position snapshot, sheet "Sua carteira"
  2. XP PosicaoDetalhada.xlsx    — position snapshot with "Fundos de Investimentos" section
  3. XP Extrato.xlsx             — account statement with fund/TD/CDB/LCA movements

Clear sheet layout ("Sua carteira"):
  Row 1:  "Conta: XXXXXXXX | DD/MM/YYYY, HH:MM"
  Sections detected dynamically by scanning for "Fundos Imobiliários" / "Ações" labels
  and sub-headers matching r'^\\d+[.,]\\d+%\\s*\\|' pattern.

  FII sub-header cols: A=ticker, F=preço_médio_abertura, H=quantidade_de_cotas
  Ação sub-header cols: A=ticker, E=preço_médio, G=qtd_total
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

logger = logging.getLogger(__name__)

# Row indices (1-based, matching openpyxl default)
_ACOES_HEADER_ROW = 8
_FIIS_HEADER_ROW = 29

# Maximum rows to scan per section before giving up
_MAX_SCAN_ROWS = 60


def _parse_brl(value: Any) -> Decimal | None:
    """Parse a BRL-formatted value like 'R$ 1.911,30' or '35,31' to Decimal."""
    if value is None:
        return None
    s = str(value).strip()
    # Remove 'R$', spaces, thousands dots; replace comma decimal separator
    s = re.sub(r"[R$\s\.]", "", s).replace(",", ".")
    if not s or s in ("-", ""):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_qty(value: Any) -> Decimal | None:
    """Parse integer quantity like '46' or '21'."""
    if value is None:
        return None
    s = str(value).strip().replace(",", "").replace(".", "")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _extract_date_from_header(ws: Any) -> date:
    """Extract transaction date from cell A1: 'Conta: 17008143 | 22/03/2026, 19:45'."""
    cell_val = ws.cell(row=1, column=1).value
    if cell_val:
        match = re.search(r"(\d{2}/\d{2}/\d{4})", str(cell_val))
        if match:
            try:
                return datetime.strptime(match.group(1), "%d/%m/%Y").date()
            except ValueError:
                pass
    return date.today()


def _find_section_header_row(ws: Any, start_row: int, expected_col1: str, tolerance: int = 5) -> int:
    """Scan rows around start_row to find the actual header row.

    Checks if column A contains a value matching expected_col1 (case-insensitive partial).
    Returns the matched row number, or start_row if not found.
    """
    for r in range(max(1, start_row - 2), start_row + tolerance + 1):
        val = ws.cell(row=r, column=1).value
        if val and expected_col1.lower() in str(val).lower():
            return r
    return start_row


def _parse_acoes_section(ws: Any, header_row: int, txn_date: date) -> list[dict]:
    """Extract ações positions starting from header_row + 1."""
    rows = []
    for r in range(header_row + 1, header_row + _MAX_SCAN_ROWS):
        ticker_val = ws.cell(row=r, column=1).value
        if ticker_val is None or str(ticker_val).strip() == "":
            break  # end of section

        ticker = str(ticker_val).strip().upper()
        if not re.match(r"^[A-Z]{4}\d{1,2}[A-Z]?$", ticker):
            # Skip non-ticker rows (subtotals, labels, etc.)
            logger.debug("xlsx_parser: skipping ações row %d — ticker='%s'", r, ticker)
            continue

        preco_medio = _parse_brl(ws.cell(row=r, column=5).value)  # col E
        quantidade = _parse_qty(ws.cell(row=r, column=7).value)    # col G

        if preco_medio is None or quantidade is None or quantidade <= 0:
            logger.warning(
                "xlsx_parser: skipping ação %s row %d — invalid preço=%s qty=%s",
                ticker, r, preco_medio, quantidade,
            )
            continue

        rows.append({
            "ticker": ticker,
            "asset_class": "acao",
            "transaction_type": "buy",
            "transaction_date": txn_date,
            "quantity": quantidade,
            "unit_price": preco_medio,
            "brokerage_fee": Decimal("0"),
            "irrf_withheld": Decimal("0"),
            "notes": "Importado de PosicaoDetalhada Clear (snapshot)",
            "parser_source": "xlsx_clear_parser",
        })
        logger.debug("xlsx_parser: ação %s qty=%s preço=%s", ticker, quantidade, preco_medio)

    return rows


def _parse_fiis_section(ws: Any, header_row: int, txn_date: date) -> list[dict]:
    """Extract FII positions starting from header_row + 1."""
    rows = []
    for r in range(header_row + 1, header_row + _MAX_SCAN_ROWS):
        ticker_val = ws.cell(row=r, column=1).value
        if ticker_val is None or str(ticker_val).strip() == "":
            break

        ticker = str(ticker_val).strip().upper()
        if not re.match(r"^[A-Z]{4}11$", ticker):
            logger.debug("xlsx_parser: skipping FII row %d — ticker='%s'", r, ticker)
            continue

        preco_medio = _parse_brl(ws.cell(row=r, column=6).value)  # col F (preço médio abertura)
        quantidade = _parse_qty(ws.cell(row=r, column=8).value)    # col H (quantidade de cotas)

        if preco_medio is None or quantidade is None or quantidade <= 0:
            logger.warning(
                "xlsx_parser: skipping FII %s row %d — invalid preço=%s qty=%s",
                ticker, r, preco_medio, quantidade,
            )
            continue

        rows.append({
            "ticker": ticker,
            "asset_class": "fii",
            "transaction_type": "buy",
            "transaction_date": txn_date,
            "quantity": quantidade,
            "unit_price": preco_medio,
            "brokerage_fee": Decimal("0"),
            "irrf_withheld": Decimal("0"),
            "notes": "Importado de PosicaoDetalhada Clear (snapshot)",
            "parser_source": "xlsx_clear_parser",
        })
        logger.debug("xlsx_parser: FII %s qty=%s preço=%s", ticker, quantidade, preco_medio)

    return rows


def _detect_rf_subtype(name: str) -> str:
    """Detect fixed income product subtype from product name.

    Returns short type code used as ticker prefix and in notes.
    Examples:
      "CDB BANCO DIGIMAIS S.A. - AGO/2030"  → "CDB"
      "LCI XP INVESTIMENTOS - JAN/2027"     → "LCI"
      "Tesouro IPCA+ 2035"                  → "NTNB"
      "Tesouro Selic 2027"                  → "LFT"
      "Tesouro Prefixado 2026"              → "LTN"
      "Debênture VALE ON"                   → "DEB"
      "SulAmérica Crédito Ativo FIRF"       → "FIF"
    """
    import unicodedata
    n = unicodedata.normalize("NFKD", name.upper()).encode("ASCII", "ignore").decode()
    if n.startswith("CDB"):
        return "CDB"
    if n.startswith("LCI"):
        return "LCI"
    if n.startswith("LCA"):
        return "LCA"
    if n.startswith("LF"):
        return "LFT"
    if n.startswith("CRI"):
        return "CRI"
    if n.startswith("CRA"):
        return "CRA"
    if "TESOURO" in n or "TREASURY" in n:
        if "SELIC" in n:
            return "LFT"
        if "IPCA" in n:
            return "NTNB"
        if "PREFIXADO" in n or "PREFIXED" in n:
            return "LTN"
        return "TD"
    if "DEBENTURE" in n or "DEBENTURA" in n:
        return "DEB"
    if any(t in n for t in ("FIF", "FIC", "FIRF", "FICFI", "FUNDO")):
        return "FIF"
    return "RF"


def _ticker_slug(name: str, prefix: str = "", max_len: int = 8) -> str:
    """Generate a synthetic ticker from a fund/product name.

    Takes uppercase alphanumeric chars from the name, skipping common stop-words.
    E.g. "SulAmérica Crédito Ativo FIRF CP LP" → "SULAMCRE"
    """
    stop = {"de", "da", "do", "das", "dos", "e", "a", "o", "fic", "fif", "fi",
            "rf", "cp", "rl", "lp", "sa", "s.a", "ltda"}
    # Normalize: remove accents/special chars, uppercase
    import unicodedata
    normalized = unicodedata.normalize("NFKD", name).encode("ASCII", "ignore").decode()
    words = re.split(r"[^A-Za-z0-9]+", normalized)
    chars = "".join(
        w[:4].upper() for w in words
        if w and w.lower() not in stop and len(w) > 1
    )
    slug = (prefix + chars)[:max_len]
    return slug or (prefix + "FUND")[:max_len]


def _parse_date_br(value: Any) -> date | None:
    """Parse Brazilian date format DD/MM/YYYY."""
    if value is None:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()
    except ValueError:
        return None


def _detect_xp_xlsx(ws: Any) -> bool:
    """Return True if this worksheet looks like an XP PosicaoDetalhada.

    XP files have "Fundos de Investimentos" or "Renda Fixa" in column A,
    whereas Clear files have "Ativo" as a column header in rows 8/29.
    """
    for r in range(1, min(20, (ws.max_row or 20) + 1)):
        val = ws.cell(row=r, column=1).value
        if val and ("fundos de investimentos" in str(val).lower()
                    or "renda fixa" in str(val).lower()):
            return True
    return False


def _parse_xp_date(ws: Any) -> date:
    """Extract date from XP header: 'Conta: 2253571 | 23/03/2026, 09:26' (row 1 col 6)."""
    for col in range(1, 14):
        val = ws.cell(row=1, column=col).value
        if val:
            m = re.search(r"(\d{2}/\d{2}/\d{4})", str(val))
            if m:
                try:
                    return datetime.strptime(m.group(1), "%d/%m/%Y").date()
                except ValueError:
                    pass
    return date.today()


def _parse_xp_fundos(ws: Any, txn_date: date) -> list[dict]:
    """Parse 'Fundos de Investimentos' section from XP XLSX.

    Section layout:
      - Section header row: col A = "Fundos de Investimentos"
      - Sub-header row:  col A = "XX% | Tipo", col B = "Posição", col F = "Valor aplicado"
      - Data rows: col A = fund name (long text), col B = position, col F = valor aplicado
      - Blank row = end of section

    We use:
      ticker     = synthetic slug from fund name
      asset_class = renda_fixa  (all RF funds in practice)
      quantity   = 1
      unit_price = Valor aplicado (col F, index 6)
    """
    rows: list[dict] = []
    in_section = False
    past_subheader = False
    seen_tickers: dict[str, int] = {}

    for r in range(1, (ws.max_row or 100) + 1):
        col_a = ws.cell(row=r, column=1).value
        a_str = str(col_a).strip() if col_a else ""

        if not in_section:
            if "fundos de investimentos" in a_str.lower():
                in_section = True
            continue

        # Detect sub-header: "XX% | ..." pattern in col A
        if not past_subheader:
            if re.match(r"^\d+[\.,]\d+%", a_str) or "posição" in str(
                ws.cell(row=r, column=2).value or ""
            ).lower():
                past_subheader = True
            continue

        # Blank row = end of this section
        if not a_str or a_str == " ":
            break

        # Skip rows that look like sub-totals or new sub-headers
        if re.match(r"^\d+[\.,]\d+%", a_str):
            continue  # another "XX% | Tipo" sub-header

        valor_aplicado = _parse_brl(ws.cell(row=r, column=6).value)
        if valor_aplicado is None or valor_aplicado <= 0:
            continue

        subtype = _detect_rf_subtype(a_str)
        base_ticker = _ticker_slug(a_str, max_len=8 - len(subtype) - 1)
        slug = f"{subtype}.{base_ticker}" if base_ticker else subtype
        seen_tickers[slug] = seen_tickers.get(slug, 0) + 1
        ticker = slug if seen_tickers[slug] == 1 else f"{slug[:6]}{seen_tickers[slug]:02d}"

        rows.append({
            "ticker": ticker,
            "asset_class": "renda_fixa",
            "transaction_type": "buy",
            "transaction_date": txn_date,
            "quantity": Decimal("1"),
            "unit_price": valor_aplicado,
            "brokerage_fee": Decimal("0"),
            "irrf_withheld": Decimal("0"),
            "notes": f"[{subtype}] XP Fundo: {a_str[:80]}",
            "parser_source": "xlsx_xp_parser",
        })
    return rows


def _parse_xp_renda_fixa(ws: Any, txn_date: date) -> list[dict]:
    """Parse 'Renda Fixa' section from XP XLSX.

    Section header row: col A = "Renda Fixa"
    Data header: col A = "XX% | ...", col B = "Posição a mercado",
                 col G = "Data aplicação", col H = "Data vencimento",
                 col I = "Quantidade", col J = "Preço Unitário"
    Data rows: col A = product name, col D = valor aplicado, col G = date,
               col I = qty (string), col J = unit price

    Col indices (1-based openpyxl):
      A=1: name, D=4: valor aplicado, G=7: data aplicação, H=8: maturity,
      I=9: quantity, J=10: unit price
    """
    rows: list[dict] = []
    in_section = False
    past_subheader = False
    seen_tickers: dict[str, int] = {}

    for r in range(1, (ws.max_row or 100) + 1):
        col_a = ws.cell(row=r, column=1).value
        a_str = str(col_a).strip() if col_a else ""

        if not in_section:
            if a_str.lower() == "renda fixa":
                in_section = True
            continue

        if not past_subheader:
            if re.match(r"^\d+[\.,]\d+%", a_str) or "posição" in str(
                ws.cell(row=r, column=2).value or ""
            ).lower():
                past_subheader = True
            continue

        if not a_str or a_str == " ":
            break

        if re.match(r"^\d+[\.,]\d+%", a_str):
            continue

        # Parse quantity and unit price
        qty_val = ws.cell(row=r, column=9).value   # col I
        price_val = ws.cell(row=r, column=10).value  # col J
        date_val = ws.cell(row=r, column=7).value    # col G (data aplicação)

        quantity = _parse_qty(qty_val) if qty_val else None
        unit_price = _parse_brl(price_val) if price_val else None

        # Fallback: use valor aplicado / quantity if unit price missing
        if unit_price is None:
            valor_apl = _parse_brl(ws.cell(row=r, column=4).value)
            if valor_apl and quantity and quantity > 0:
                unit_price = valor_apl / quantity
            elif valor_apl:
                quantity = Decimal("1")
                unit_price = valor_apl

        if unit_price is None or unit_price <= 0:
            continue
        if quantity is None or quantity <= 0:
            quantity = Decimal("1")

        txn_date_row = _parse_date_br(date_val) or txn_date

        subtype = _detect_rf_subtype(a_str)
        base_ticker = _ticker_slug(a_str, max_len=8 - len(subtype) - 1)
        slug = f"{subtype}.{base_ticker}" if base_ticker else subtype
        seen_tickers[slug] = seen_tickers.get(slug, 0) + 1
        ticker = slug if seen_tickers[slug] == 1 else f"{slug[:6]}{seen_tickers[slug]:02d}"

        rows.append({
            "ticker": ticker,
            "asset_class": "renda_fixa",
            "transaction_type": "buy",
            "transaction_date": txn_date_row,
            "quantity": quantity,
            "unit_price": unit_price,
            "brokerage_fee": Decimal("0"),
            "irrf_withheld": Decimal("0"),
            "notes": f"[{subtype}] XP RF: {a_str[:80]}",
            "parser_source": "xlsx_xp_parser",
        })
    return rows


def parse_xp_xlsx(xlsx_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Parse XP PosicaoDetalhada.xlsx and return synthetic buy transactions.

    Handles two sections: Fundos de Investimentos and Renda Fixa.
    """
    try:
        import openpyxl
    except ImportError:
        return [], ["openpyxl not installed"]

    warnings: list[str] = []
    transactions: list[dict] = []

    try:
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"Failed to open XLSX: {exc}"]

    ws = wb.active
    if ws is None:
        return [], ["XLSX has no sheets"]

    txn_date = _parse_xp_date(ws)
    logger.info("xlsx_xp_parser: reference date %s", txn_date)

    fundos = _parse_xp_fundos(ws, txn_date)
    if fundos:
        transactions.extend(fundos)
        logger.info("xlsx_xp_parser: parsed %d fundos", len(fundos))
    else:
        warnings.append("Nenhum fundo encontrado na seção Fundos de Investimentos")

    rf = _parse_xp_renda_fixa(ws, txn_date)
    if rf:
        transactions.extend(rf)
        logger.info("xlsx_xp_parser: parsed %d renda fixa", len(rf))
    else:
        warnings.append("Nenhum produto encontrado na seção Renda Fixa")

    if not transactions:
        warnings.append("Nenhuma posição encontrada. Verifique se é um PosicaoDetalhada.xlsx da XP.")

    wb.close()
    return transactions, warnings


def parse_xlsx(xlsx_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Auto-detect broker format and dispatch to correct parser.

    Detects XP vs Clear by scanning for XP-specific section headers.
    Falls back to Clear parser if format is unrecognized.
    """
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb.active
        is_xp = _detect_xp_xlsx(ws) if ws else False
        wb.close()
    except Exception:
        is_xp = False

    if is_xp:
        logger.info("xlsx_parser: detected XP PosicaoDetalhada format")
        return parse_xp_xlsx(xlsx_bytes)

    # Check for XP Extrato (account statement) format
    try:
        from io import BytesIO as _BytesIO
        import openpyxl as _opx
        wb2 = _opx.load_workbook(_BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws2 = wb2.active
        is_xp_extrato = _detect_xp_extrato(ws2) if ws2 else False
        wb2.close()
    except Exception:
        is_xp_extrato = False

    if is_xp_extrato:
        logger.info("xlsx_parser: detected XP Extrato format")
        return parse_xp_extrato_xlsx(xlsx_bytes)

    logger.info("xlsx_parser: detected Clear format (or unknown → fallback)")
    return parse_clear_xlsx(xlsx_bytes)


def parse_clear_xlsx(xlsx_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Parse Clear PosicaoDetalhada.xlsx and return synthetic buy transactions.

    Section detection is done dynamically by scanning for main section labels
    ('Fundos Imobiliários', 'Ações') and sub-header rows matching 'XX% | Label'.
    This handles layout variations across different Clear export versions.

    Column layout (1-based, openpyxl):
      FII sub-header: A=ticker, F=preço_médio_abertura, H=quantidade_de_cotas
      Ação sub-header: A=ticker, E=preço_médio, G=qtd_total
    """
    try:
        import openpyxl
    except ImportError:
        return [], ["openpyxl not installed — cannot parse XLSX files"]

    warnings: list[str] = []
    transactions: list[dict] = []

    try:
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"Failed to open XLSX: {exc}"]

    # Find the main sheet ("Sua carteira" or first sheet)
    ws = None
    for name in wb.sheetnames:
        if "carteira" in name.lower() or "posicao" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    if ws is None:
        return [], ["XLSX has no sheets"]

    # Extract reference date from header (row 1)
    txn_date = _extract_date_from_header(ws)
    logger.info("xlsx_parser: extracted date %s", txn_date)

    # Scan all rows to find section context and sub-headers
    # Clear layout uses:
    #   - Section label rows: "Fundos Imobiliários", "Ações" in col A
    #   - Sub-header rows:    "XX,X% | Fundos Listados", "XX,X% | Renda Variável Brasil" in col A
    # We track which main section we're in, then on seeing a sub-header, we parse data.
    current_section: str | None = None  # "fii" | "acao"
    parsed_fii = False
    parsed_acao = False

    _subheader_re = re.compile(r"^\d+[\.,]\d+%\s*\|", re.UNICODE)

    max_row = ws.max_row or 200
    for r in range(1, max_row + 1):
        col_a = ws.cell(row=r, column=1).value
        a_str = str(col_a).strip() if col_a not in (None, " ", "") else ""

        if not a_str:
            continue

        # Detect main section labels
        a_lower = a_str.lower()
        import unicodedata as _ud
        a_norm = _ud.normalize("NFKD", a_lower).encode("ascii", "ignore").decode()

        if "fundos imobiliarios" in a_norm or ("fundo" in a_norm and "imobili" in a_norm):
            current_section = "fii"
            continue
        if a_norm.startswith("acoes") or a_norm == "acoes" or (
            "acoes" in a_norm and len(a_norm) < 10
        ):
            current_section = "acao"
            continue

        # Detect sub-header rows (e.g. "45,1% | Fundos Listados")
        if _subheader_re.match(a_str):
            if current_section == "fii" and not parsed_fii:
                fiis = _parse_fiis_section(ws, r, txn_date)
                if fiis:
                    transactions.extend(fiis)
                    logger.info("xlsx_parser: parsed %d FIIs from row %d", len(fiis), r)
                    parsed_fii = True
            elif current_section == "acao" and not parsed_acao:
                acoes = _parse_acoes_section(ws, r, txn_date)
                if acoes:
                    transactions.extend(acoes)
                    logger.info("xlsx_parser: parsed %d ações from row %d", len(acoes), r)
                    parsed_acao = True

        if parsed_fii and parsed_acao:
            break  # got both sections, done

    if not parsed_fii:
        warnings.append("Nenhum FII encontrado na seção Fundos Imobiliários")
    if not parsed_acao:
        warnings.append("Nenhuma ação encontrada na seção Ações")
    if not transactions:
        warnings.append(
            "Nenhuma posição encontrada. Verifique se o arquivo é um PosicaoDetalhada.xlsx da Clear."
        )

    wb.close()
    return transactions, warnings


# ---------------------------------------------------------------------------
# XP Extrato parser (account statement: funds, TD, CDB, LCA)
# ---------------------------------------------------------------------------

def _detect_xp_extrato(ws: Any) -> bool:
    """Return True if the sheet looks like an XP account extrato."""
    for r in range(1, 20):
        for c in range(1, 14):
            val = ws.cell(row=r, column=c).value
            if val and "extrato da conta" in str(val).lower():
                return True
    return False


def _extract_fund_ticker(description: str) -> str:
    """Extract a short slug from a fund/instrument description for use as ticker."""
    import unicodedata
    d = description.upper()
    d_norm = unicodedata.normalize("NFKD", d).encode("ascii", "ignore").decode()

    if "TESOURO DIRETO" in d_norm:
        return "TESOURO"

    # Remove action prefix words
    for prefix in [
        r"^RESGATE\s+", r"^COMPRA\s+", r"^APLICACAO\s+", r"^LIQUIDACAO\s+",
        r"^TED\s+APLICACAO\s+FUNDOS\s+", r"^IRRF\s+S/RESGATE\s+FUNDOS\s+-\s+",
        r"^IOF\s+S/\s+RESGATE\s+FUNDOS\s+-\s+", r"^\w+\s*\|\s*",
    ]:
        d_norm = re.sub(prefix, "", d_norm, count=1)

    # Strip product type prefix
    for ptype in ("CDB BANCO ", "LCA BANCO ", "LCI BANCO ", "CDB ", "LCA ", "LCI "):
        if d_norm.startswith(ptype):
            d_norm = d_norm[len(ptype):]
            break

    # Take first 2 words, max 8 chars
    words = re.split(r"[\s\-/]+", d_norm.strip())
    slug = "".join(w[:4] for w in words[:2])
    slug = re.sub(r"[^A-Z0-9]", "", slug)[:8]
    return slug or "FUNDO"


def parse_xp_extrato_xlsx(xlsx_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Parse XP account extrato XLSX (fund/TD/CDB/LCA movements).

    Column layout (1-based):
      B=data_movimentacao, C=data_liquidacao, D=lancamento, F=valor_rs, G=saldo

    Mapped events:
      COMPRA TESOURO DIRETO           → buy  tesouro_direto  TESOURO
      LIQUIDACAO TESOURO DIRETO       → sell tesouro_direto  TESOURO
      COMPRA CDB/LCA/LCI              → buy  renda_fixa      <slug>
      RESGATE ... | LCA/CDB/LCI/CRI   → sell renda_fixa      <slug>
      TED ... APLICACAO FUNDOS <name> → buy  fundo_investimento <slug>
      RESGATE <fund> (not LCA/CDB)    → sell fundo_investimento <slug>
      IRRF, IOF, TEDs simples, TED    → skipped
    """
    try:
        import openpyxl
        from io import BytesIO
        from datetime import datetime as _dt
    except ImportError:
        return [], ["openpyxl not installed"]

    try:
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"Failed to open XLSX: {exc}"]

    ws = wb.active
    if ws is None:
        return [], ["XLSX has no sheets"]

    import unicodedata
    warnings_list: list[str] = []
    transactions: list[dict] = []
    skipped = 0

    # Find data start row (look for row with "Movimentação"/"Liquidação" headers)
    data_start = None
    for r in range(1, 30):
        val_b = str(ws.cell(row=r, column=2).value or "").lower()
        val_c = str(ws.cell(row=r, column=3).value or "").lower()
        if "movimenta" in val_b or "liquida" in val_c:
            data_start = r + 1
            break

    if data_start is None:
        return [], ["Cabeçalho de dados não encontrado no extrato XP"]

    _SKIP = ["IRRF S/", "IOF S/", "TRANSFERENCIA RECEBIDA", "TRANSFERENCIA ENVIADA",
             "TED BCO ", "RETIRADA EM C/C", "RECEBIMENTO DE TED",
             "LANCAMENTOS FUTUROS", "NAO HA LANCAMENTOS"]

    for r in range(data_start, (ws.max_row or 300) + 1):
        date_val = ws.cell(row=r, column=2).value
        desc_val = ws.cell(row=r, column=4).value
        valor_val = ws.cell(row=r, column=6).value

        if not isinstance(date_val, _dt):
            continue
        if desc_val is None or valor_val is None:
            continue

        desc = str(desc_val).strip()
        dn = unicodedata.normalize("NFKD", desc.upper()).encode("ascii", "ignore").decode()

        try:
            valor = float(valor_val)
        except (TypeError, ValueError):
            continue
        if valor == 0:
            continue

        txn_date = date_val.date()

        # Check skip list (but allow TED APLICACAO FUNDOS through)
        if any(p in dn for p in _SKIP) and "TED APLICACAO FUNDOS" not in dn:
            skipped += 1
            continue

        asset_class = txn_type = ticker = notes_extra = None

        if "TESOURO DIRETO" in dn:
            asset_class, ticker = "tesouro_direto", "TESOURO"
            if "COMPRA" in dn:
                txn_type = "buy"
            elif "LIQUIDACAO" in dn or "VENCIMENTO" in dn:
                txn_type = "sell"

        elif dn.startswith("COMPRA CDB") or dn.startswith("COMPRA LCA") or dn.startswith("COMPRA LCI"):
            asset_class, txn_type = "renda_fixa", "buy"
            ticker = _extract_fund_ticker(desc)

        elif "RESGATE" in dn and any(p in dn for p in ("| LCA", "| LCI", "| CDB", "| CRI", "| CRA")):
            asset_class, txn_type = "renda_fixa", "sell"
            ticker = _extract_fund_ticker(desc)

        elif "TED APLICACAO FUNDOS" in dn:
            asset_class, txn_type = "fundo_investimento", "buy"
            m = re.search(r"TED\s+APLICA\S*\s+FUNDOS\s+(.+)", dn)
            fund_name = m.group(1).strip() if m else desc
            ticker = _extract_fund_ticker(fund_name)
            notes_extra = fund_name[:50]

        elif dn.startswith("RESGATE ") and not any(p in dn for p in ("| LCA", "| LCI", "| CDB", "TESOURO")):
            asset_class, txn_type = "fundo_investimento", "sell"
            ticker = _extract_fund_ticker(desc[8:])
            notes_extra = desc[8:60]

        if not (ticker and asset_class and txn_type):
            skipped += 1
            continue

        amount = Decimal(str(abs(valor))).quantize(Decimal("0.01"))
        note = f"XP Extrato {txn_date.strftime('%d/%m/%Y')}: {desc[:70]}"
        if notes_extra:
            note += f" [{notes_extra}]"

        transactions.append({
            "ticker": ticker[:12],
            "asset_class": asset_class,
            "transaction_type": txn_type,
            "transaction_date": txn_date,
            "quantity": Decimal("1"),
            "unit_price": amount,
            "brokerage_fee": Decimal("0"),
            "irrf_withheld": Decimal("0"),
            "notes": note,
            "parser_source": "xp_extrato_xlsx",
        })

    wb.close()
    logger.info("xp_extrato_parser: %d transactions, %d skipped", len(transactions), skipped)
    if not transactions:
        warnings_list.append("Nenhuma transação encontrada no extrato XP.")
    return transactions, warnings_list
